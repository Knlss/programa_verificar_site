from concurrent.futures import ThreadPoolExecutor
import file_utils as file
from openpyxl.utils import column_index_from_string
import temporario.config.settings as settings
from openpyxl.styles import Font, Alignment, Border
import requests

def check_access(site):
    try:
        response = requests.get(site, verify=False, timeout=10) 
        return "Accessible" if response.status_code == 200 else "Inaccessible"
    except requests.exceptions.RequestException as e:
        if isinstance(e, requests.exceptions.Timeout):
            return "Timeout Exceeded"
        else:
            return "Inaccessible"

def verif_columns_entry(input, output, sites):
    if input is not None:
        if output is not None:
            if len(input) == 1 and 'A' <= input.upper() <= 'Z':
                if len(output) == 1 and 'A' <= output.upper() <= 'Z' and output.upper() != input.upper():
                    if sites:
                        return True
                    else: return "MENSAGEM DE ERRO"
                else: return "MENSAGEM DE ERRO"
            else: return "MENSAGEM DE ERRO"
        else: return "MENSAGEM DE ERRO"
    else: return "MENSAGEM DE ERRO"

def calculate_verified_sites(acess, inacess, timeout):
    verified = acess, inacess, timeout
    settings.verified_sites = verified

def column_letter_converter(input_value, output_value):
    input = input_value.upper()
    output = output_value.upper()
    set_column_value(input, output)

def set_column_value(input, output):
    settings.input_column_value = input
    settings.output_column_value = output

def verif_cells(ws, column_value):
    sites = []
    for cell in ws[column_value]:
        if cell.value is not None:
            if cell.value.startswith("http://"):
                sites.append(cell.value)
            else:
                sites.append("http://" + cell.value)
    return sites

def write_verification_results():
    sites = verif_cells(file.ws, settings.input_column_value)
    column_output = column_index_from_string(settings.output_column_value)
    with ThreadPoolExecutor(max_workers=10) as executor: 
        results = list(executor.map(check_access, sites))
                                
        # Adicionar os resultados na coluna E (ou 5) alinhados com os respectivos sites na coluna B
        for i, result in enumerate(results, start=1):
            if i <= len(sites):
                file.ws.cell(row=i, column=column_output).value = result
                file.ws.cell(row=i, column=column_output).alignment = settings.alignment_cell
                                        
                # Atualizar a cor da fonte com base no resultado
                if result == "Accessible":
                    settings.accessible_sites += 1
                    file.ws.cell(row=i, column=column_output).font = Font(size=settings.font_size_e, color=settings.font_color_acessible)
                elif result == "Inaccessible":
                    settings.inaccessible_sites += 1
                    file.ws.cell(row=i, column=column_output).font = Font(size=settings.font_size_e, color=settings.font_color_inacessible)
                elif result == "Timeout Exceeded":
                    settings.timeout_sites += 1 
                    file.ws.cell(row=i, column=column_output).font = Font(size=settings.font_size_e, color=settings.font_color_timeout)

def config_output_first_cell(ws, output_column):
    ws[f'{output_column}1'].value = settings.cell_e1_text
    ws[f'{output_column}1'].font = Font(bold=True, size=settings.first_cell_size)
    ws[f'{output_column}1'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{output_column}1'].border = Border(left=settings.title_border_left, right=settings.title_border_right, top=settings.title_border_top, bottom=settings.title_border_bottom)

def apply_middle_borders(ws, output_column):
    for i in range(2, settings.verified_sites + 1):
        cell = ws.cell(row=i, column=column_index_from_string(output_column))
        cell.border = Border(left=settings.cell_border_left, right=settings.cell_border_right, top=settings.cell_border_top, bottom=settings.cell_border_bottom)

def apply_upper_border(ws, output_column):
    e2_cell = ws[f'{output_column}2']
    e2_cell.border = Border(left=settings.cell_border_left, right=settings.cell_border_right, top=settings.end_border_edges_top, bottom=settings.cell_border_bottom)

def apply_lower_border(ws, output_column):
    ex_cell = ws[f'{output_column}{settings.verified_sites + 1}']
    ex_cell.border = Border(left=settings.cell_border_left, right=settings.cell_border_right, top=settings.cell_border_top, bottom=settings.end_border_edges_bottom)

def set_column_width(ws, input_column, output_column):
    input_column_width = ws.column_dimensions[input_column].width
    ws.column_dimensions[output_column].width = input_column_width
