import tkinter as tk
from tkinter import filedialog
import os
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
import warnings
from urllib3.exceptions import InsecureRequestWarning

# Ignore os avisos sobre solicitações HTTP não verificadas
warnings.filterwarnings("ignore", category=InsecureRequestWarning)

# Inicializar a variável global timeout_sites
column_value = None

def check_access(site):
    try:
        response = requests.get(site, verify=False, timeout=10)  # Ignorar a verificação SSL e definir um timeout de 10 segundos
        return "Accessible" if response.status_code == 200 else "Inaccessible"
    except requests.exceptions.RequestException as e:
        if isinstance(e, requests.exceptions.Timeout):
            return "Timeout Exceeded"
        else:
            return "Inaccessible"

def save_value():
    global column_value
    column_value = entry.get()

def process_file():
    border_color = '000000'  # Cor em formato hexadecimal (por exemplo, preto)
    font_color_acessible = '364624'  # Cor preta
    font_color_inacessible = '083825'  # Cor vermelha
    font_color_timeout = '0000FF'  # Cor azul
    border_left = Side(style='thin', color=border_color)  # Borda fina preta no lado esquerdo
    border_right = Side(style='thin', color=border_color)  # Borda fina preta no lado direito
    border_top = Side(style='thin', color=border_color)  # Borda fina preta na parte superior
    border_bottom = Side(style='thin', color=border_color)  # Borda fina preta na parte inferior
    border_bottom_none = None
    border_top_none = None
    global column_value
    timeout_sites = 0
    accessible_sites = 0  # Definir a variável accessible_sites com um valor inicial
    inaccessible_sites = -1  # Definir a variável inaccessible_sites com um valor inicial
    verified_sites = 0  # Definir a variável verified_sites com um valor inicial
    # Solicitar ao usuário que selecione o arquivo Excel
    filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])
    if filename:
        # Carregar o arquivo Excel e selecionar a planilha ativa
        wb = load_workbook(filename)
        ws = wb.active

        # Definir o tamanho da fonte para as células na coluna E
        font_size_e = 11

        # Definir o texto para escrever na célula E1
        cell_e1_text = "COLUMN B - STATUS"

        # Definir o nome do novo arquivo Excel que será criado
        output_filename = os.path.join(os.path.expanduser("~/Documents/output_folder"), os.path.splitext(os.path.basename(filename))[0] + '_result2.xlsx')

        # Verificar se column_value está definido
        if column_value is not None:
            # Verificar se column_value é uma letra de coluna válida
            if len(column_value) == 1 and 'A' <= column_value.upper() <= 'Z':
                column_letter = column_value.upper()
                # Obter os valores de todas as células da coluna especificada e adicionar "http://" se necessário
                sites = []
                for cell in ws[column_letter]:
                    if cell.value is not None:
                        if cell.value.startswith("http://"):
                            sites.append(cell.value)
                        else:
                            sites.append("http://" + cell.value)
        
                # Verificar se há sites antes de prosseguir
                if sites:
                    
                    # Verificar os sites em paralelo
                    with ThreadPoolExecutor(max_workers=10) as executor:  # Executar até 10 verificações simultaneamente
                        results = list(executor.map(check_access, sites))
                        
                        # Adicionar os resultados na coluna E (ou 5) alinhados com os respectivos sites na coluna B
                        for i, result in enumerate(results, start=1):
                            if i <= ws.max_row:
                                ws.cell(row=i, column=5).value = result
                                
                                # Atualizar a cor da fonte com base no resultado
                                if result == "Accessible":
                                    accessible_sites += 1
                                    ws.cell(row=i, column=5).font = Font(size=font_size_e, color=font_color_acessible)
                                elif result == "Inaccessible":
                                    inaccessible_sites += 1
                                    ws.cell(row=i, column=5).font = Font(size=font_size_e, color=font_color_inacessible)
                                elif result == "Timeout Exceeded":
                                    timeout_sites += 1  # Incrementar timeout_sites se o tempo limite for excedido
                                    ws.cell(row=i, column=5).font = Font(size=font_size_e, color=font_color_timeout)

                    # Calcular o número total de sites verificados
                    verified_sites = accessible_sites + inaccessible_sites + timeout_sites

                    # Salvar o arquivo com os resultados
                    wb.save(output_filename)
                    status_label.config(text=f"Results saved in '{output_filename}'.")

                    # Escrever o texto definido na célula E1
                    ws['E1'] = cell_e1_text
                    ws['E1'].alignment = Alignment(horizontal='center', vertical='center')

                    # Definir o tamanho da fonte para as células na coluna E
                    for i in range(2, verified_sites + 2):
                        cell = ws.cell(row=i, column=5)
                        cell.font = Font(size=font_size_e)

                    # Aplicar bordas às células na coluna E
                    for i in range(2, verified_sites + 1):
                        cell = ws.cell(row=i, column=5)
                        cell.border = Border(left=border_left, right=border_right, top=border_top_none, bottom=border_bottom_none)
                        
                    e2_cell = ws['E2']
                    e2_cell.border = Border(left=border_left, right=border_right, top=border_top, bottom=border_bottom_none)
                    ex_cell = ws[f'E{verified_sites + 1}']
                    ex_cell.border = Border(left=border_left, right=border_right, top=border_top_none, bottom=border_bottom)

                    # Definir a largura da coluna E
                    largura_coluna_B = ws.column_dimensions[column_value.upper()].width
                    column_letter = get_column_letter(5)  # 5 é o índice da coluna E
                    ws.column_dimensions[column_letter].width = largura_coluna_B
                    wb.save(output_filename)
                else:
                    status_label.config(text=f"No sites found in column {column_value}.")
            else:
                status_label.config(text="Invalid column value. Please enter a letter from A to Z.")
        else:
            status_label.config(text="No column value entered. Please enter a column value first.")

# Criar a janela principal
root = tk.Tk()
root.title("Site Verifier")

# Criar um botão para abrir o arquivo Excel
open_button = tk.Button(root, text="Open Excel File", command=process_file)
open_button.pack(pady=10)

# Campo de entrada para aceitar um caractere
entry_frame = tk.Frame(root)
entry_frame.pack(pady=10)

entry_label = tk.Label(entry_frame, text="Enter a column letter (A-Z):")
entry_label.pack(side=tk.LEFT)

entry = tk.Entry(entry_frame, width=5)
entry.pack(side=tk.LEFT)
entry.focus_set()

# Botão para salvar o caractere
save_button = tk.Button(entry_frame, text="Save", command=save_value)
save_button.pack(side=tk.LEFT)

# Rótulo para exibir o status
status_label = tk.Label(root, text="")
status_label.pack()

# Iniciar o loop de eventos
root.mainloop()
