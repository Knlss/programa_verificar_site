import tkinter as tk
from tkinter import filedialog
import os
import requests
import warnings
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning
from concurrent.futures import ThreadPoolExecutor

# Ignore os avisos sobre solicitações HTTP não verificadas
warnings.filterwarnings("ignore", category=InsecureRequestWarning)

def verificar_acesso(site):
    try:
        resposta = requests.get(site, verify=False, timeout=10)  # Ignorar a verificação SSL e definir um timeout de 10 segundos
        return "Acessível" if resposta.status_code == 200 else "Inacessível"
    except requests.exceptions.RequestException as e:
        if isinstance(e, requests.exceptions.Timeout):
            return "Tempo limite excedido"
        else:
            return "Inacessível"

def processar_arquivo():
    filename = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")])
    if filename:
        wb = load_workbook(filename)
        ws = wb.active
        
        # Obter os valores de todas as células da coluna B e adicionar "http://" se necessário
        sites = []
        for cell in ws['B']:
            if cell.value is not None:
                if cell.value.startswith("http://"):
                    sites.append(cell.value)
                else:
                    sites.append("http://" + cell.value)
        
        # Verificar se há sites antes de prosseguir
        if sites:
            # Verificar os sites em paralelo
            with ThreadPoolExecutor(max_workers=10) as executor:  # Executar até 10 verificações simultaneamente
                resultados = list(executor.map(verificar_acesso, sites))
            
            # Adicionar os resultados na coluna E (ou 5)
            for i, resultado in enumerate(resultados, start=1):
                ws.cell(row=i, column=5).value = resultado
                
            # Criar uma pasta para salvar o arquivo de saída dentro do diretório 'Documents'
            output_folder = os.path.expanduser("~/Documents/output_folder")
            if not os.path.exists(output_folder):
                os.makedirs(output_folder)
            
            # Salvar o arquivo com os resultados na pasta criada
            output_filename = os.path.join(output_folder, os.path.splitext(os.path.basename(filename))[0] + '_resultado.xlsx')
            wb.save(output_filename)
            status_label.config(text=f"Resultados salvos em '{output_filename}'.")
        else:
            status_label.config(text="Nenhum site encontrado na coluna B.")

# Criar a janela principal
root = tk.Tk()
root.title("Verificador de Sites")

# Criar um botão para abrir o arquivo Excel
open_button = tk.Button(root, text="Abrir Arquivo Excel", command=processar_arquivo)
open_button.pack(pady=10)

# Rótulo para exibir o status
status_label = tk.Label(root, text="")
status_label.pack()

# Iniciar o loop de eventos
root.mainloop()