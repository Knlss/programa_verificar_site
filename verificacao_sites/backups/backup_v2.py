import tkinter as tk
from tkinter import filedialog
import os
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import requests
import warnings
from urllib3.exceptions import InsecureRequestWarning

# Ignore os avisos sobre solicitações HTTP não verificadas
warnings.filterwarnings("ignore", category=InsecureRequestWarning)

def verificar_acesso(site):
    try:
        resposta = requests.get(site, verify=False, timeout=10)  # Ignorar a verificação SSL e definir um timeout de 10 segundos
        return "Acessível" if resposta.status_code == 200 else "Inacessível"
    except requests.exceptions.RequestException as e:
        if isinstance(e, requests.exceptions.Timeout):
            return "Tempo limite excedido"
        else:
            return "Inacessível"

def processar_arquivo():

    # Declarar variáveis
    sites_acessiveis = 0
    sites_inacessiveis = 0
    sites_verificados = 0 
    border_color = '000000'  # Cor em formato hexadecimal (por exemplo, preto)
    border_left = Side(style='thin', color=border_color)  # Borda fina preta no lado esquerdo
    border_right = Side(style='thin', color=border_color)  # Borda fina preta no lado direito
    border_top = Side(style='thin', color=border_color)  # Borda fina preta na parte superior
    border_bottom = Side(style='thin', color=border_color)  # Borda fina preta na parte inferior
    border_bottom_none = None
    border_top_none = None

    # Solicitar ao usuário que selecione o arquivo Excel
    filename = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")])
    if filename:
        # Carregar o arquivo Excel e selecionar a planilha ativa
        wb = load_workbook(filename)
        ws = wb.active
        
        # Obter os valores de todas as células da coluna B e adicionar "http://" se necessário
        sites = []
        for cell in ws['B']:
            if cell.value is not None:
                if cell.value.startswith("http://"):
                    sites.append(cell.value)
                else:
                    sites.append("http://" + cell.value)
        
        # Verificar se há sites antes de prosseguir
        if sites:
            
            # Verificar os sites em paralelo
            with ThreadPoolExecutor(max_workers=10) as executor:  # Executar até 10 verificações simultaneamente
                resultados = list(executor.map(verificar_acesso, sites))
                        
                # Adicionar os resultados na coluna E (ou 5) alinhados com os respectivos sites na coluna B
                for i, resultado in enumerate(resultados, start=1):
                    ws.cell(row=i, column=5).value = resultado
                    
                    # Incrementar a contagem de sites acessíveis e inacessíveis
                    if resultado == "Acessível":
                        sites_acessiveis += 1
                    else:
                        sites_inacessiveis += 1

            # Calcular o número total de sites verificados
            sites_verificados = sites_acessiveis + sites_inacessiveis
            
            # Adicionar os resultados na coluna E (ou 5) alinhados com os respectivos sites na coluna B
            for i, resultado in enumerate(resultados, start=1):
                ws.cell(row=i, column=5).value = resultado

            # Criar uma pasta para salvar o arquivo de saída dentro do diretório 'Documents'
            output_folder = os.path.expanduser("~/Documents/output_folder")
            if not os.path.exists(output_folder):
                os.makedirs(output_folder)
            
            # Salvar o arquivo com os resultados na pasta criada
            output_filename = os.path.join(output_folder, os.path.splitext(os.path.basename(filename))[0] + '_resultado.xlsx')
            wb.save(output_filename)
            status_label.config(text=f"Resultados salvos em '{output_filename}'.")

            # Escrever "COLUNA B - STATUS" na célula E1 depois de salvar o arquivo
            ws['E1'] = "COLUNA B - STATUS"
            e1_cell = ws['E1']
            e1_cell.font = Font(bold=True, size=11)

            # Definir o tamanho da fonte para as X primeiras células na coluna E
            for i in range(2, sites_verificados + 1):
                cell = ws.cell(row=i, column=5)  # Acesso à célula na coluna E
                cell.font = Font(size=11)  # Definir o tamanho da fonte como 11

            # Aplicar bordas às células na coluna E
            for i in range(2, sites_verificados):
                cell = ws.cell(row=i, column=5)  # Acesso à célula na coluna E
                cell.border = Border(left=border_left, right=border_right, top=border_top_none, bottom=border_bottom_none)
                
            e2_cell = ws['E2']
            e2_cell.border = Border(left=border_left, right=border_right, top=border_top, bottom=border_bottom_none)
            ex_cell = ws[f'E{sites_verificados}']
            ex_cell.border = Border(left=border_left, right=border_right, top=border_top_none, bottom=border_bottom)

            # Centralizar o texto na célula E1
            ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
            largura_coluna_B = ws.column_dimensions['B'].width
            column_letter = get_column_letter(5)  # 5 é o índice da coluna E
            ws.column_dimensions[column_letter].width = largura_coluna_B
            wb.save(output_filename)
        else:
            status_label.config(text="Nenhum site encontrado na coluna B.")

# Criar a janela principal
root = tk.Tk()
root.title("Verificador de Sites")

# Criar um botão para abrir o arquivo Excel
open_button = tk.Button(root, text="Abrir Arquivo Excel", command=processar_arquivo)
open_button.pack(pady=10)

# Rótulo para exibir o status
status_label = tk.Label(root, text="")
status_label.pack()

# Iniciar o loop de eventos
root.mainloop()
