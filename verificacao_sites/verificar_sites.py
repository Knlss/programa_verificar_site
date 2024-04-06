import tkinter as tk
from tkinter import filedialog
import os
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import column_index_from_string
import requests
import warnings
from urllib3.exceptions import InsecureRequestWarning

# Ignore os avisos sobre solicitações HTTP não verificadas
warnings.filterwarnings("ignore", category=InsecureRequestWarning)

# Inicializar a variável global timeout_sites
input_column_value = None
output_column_value = None
font_size_e = None
border_color = '000000'  # Cor em formato hexadecimal (por exemplo, preto)
font_color_acessible = '4B0082'  # Cor preta
font_color_inacessible = '006400'  # Cor vermelha
font_color_timeout = '00008B'  # Cor azul
cell_border_left = None
cell_border_right = None
cell_border_top = None
cell_border_bottom = None
end_border_edges_top = None
end_border_edges_bottom = None
title_border_left = None
title_border_right = None
title_border_top = None
title_border_bottom = None
timeout_sites = 0
accessible_sites = 0  # Definir a variável accessible_sites com um valor inicial
inaccessible_sites = -1  # Definir a variável inaccessible_sites com um valor inicial
verified_sites = 0  # Definir a variável verified_sites com um valor inicial
alignment_cell = Alignment(horizontal='center', vertical='center')
cell_e1_text = "COLUMN B - STATUS"

root = tk.Tk()
cell_border_left_config = tk.BooleanVar()
cell_border_right_config = tk.BooleanVar()
cell_border_top_config = tk.BooleanVar()
cell_border_bottom_config = tk.BooleanVar()
title_border_left_config = tk.BooleanVar()
title_border_right_config = tk.BooleanVar()
title_border_top_config = tk.BooleanVar()
title_border_bottom_config = tk.BooleanVar()
end_border_edges_config = tk.BooleanVar()

# Campo de entrada para aceitar tamanho da fonte
entry_cell_font = tk.Frame(root)
entry_cell_font.pack(pady=5)
entry_font_label = tk.Label(entry_cell_font, text="Enter a font size (8-18):")
entry_font_label.pack(side=tk.LEFT)
entry_font = tk.Entry(entry_cell_font, width=5)
entry_font.pack(side=tk.LEFT)

# Campo de entrada para aceitar nome da célula 1
entry_title_name = tk.Frame(root)
entry_title_name.pack(pady=5)
entry_title_label = tk.Label(entry_title_name, text="Enter a title name (max 15 chars):")
entry_title_label.pack(side=tk.LEFT)


def check_font(font):
    if int(font) >= 8 and int(font) <= 18:
        return True
    else:
        return False

def validate_entry(text):
    return len(text) <= 15
    
validate_cmd = root.register(validate_entry)
entry_title = tk.Entry(entry_title_name, validate="key", width=20, validatecommand=(validate_cmd, '%P'))
entry_title.pack(side=tk.LEFT)

def root_click(event):
    if event.num == 1:  # Verifica se o evento é um clique esquerdo do mouse
        if not isinstance(event.widget, tk.Button) and not isinstance(event.widget.master, tk.Frame):
            menu_cores.pack_forget()
            print(font_size_e)

def check_access(site):
    try:
        response = requests.get(site, verify=False, timeout=10)  # Ignorar a verificação SSL e definir um timeout de 10 segundos
        return "Accessible" if response.status_code == 200 else "Inaccessible"
    except requests.exceptions.RequestException as e:
        if isinstance(e, requests.exceptions.Timeout):
            return "Timeout Exceeded"
        else:
            return "Inaccessible"
        
def default_return():

    global font_size_e, border_color, font_color_acessible, font_color_inacessible, font_color_timeout, cell_border_left, cell_border_right, cell_border_top, cell_border_bottom, end_border_edges_top, end_border_edges_bottom, title_border_left, title_border_right, title_border_top, title_border_bottom, alignment_cell, cell_e1_text

    if return_default_value.get():
        checkbox1.config(state="disabled")
        checkbox2.config(state="disabled")
        checkbox3.config(state="disabled")
        checkbox4.config(state="disabled")
        checkbox5.config(state="disabled")
        checkbox6.config(state="disabled")
        checkbox7.config(state="disabled")
        checkbox8.config(state="disabled")
        checkbox9.config(state="disabled")


        entry_font.delete(0, "end")  # Limpa o Entry
        entry_font.insert(0, 11)  # Define o texto padrão
        entry_font.config(state="disabled")

        cell_border_right_config.set(False)
        cell_border_left_config.set(False)
        cell_border_top_config.set(False)
        cell_border_bottom_config.set(False)
        title_border_right_config.set(False)
        title_border_left_config.set(False)
        title_border_top_config.set(False)
        title_border_bottom_config.set(False)
        end_border_edges_config.set(True)

        font_size_e = 11
        border_color = '000000'  # Cor em formato hexadecimal (por exemplo, preto)
        font_color_acessible = '4B0082'  # Cor preta
        font_color_inacessible = '006400'  # Cor vermelha
        font_color_timeout = '00008B'  # Cor azul
        cell_border_left = None
        cell_border_right = None
        cell_border_top = None
        cell_border_bottom = None
        end_border_edges_top = None
        end_border_edges_bottom = None
        title_border_left = None
        title_border_right = None
        title_border_top = None
        title_border_bottom = None
        alignment_cell = Alignment(horizontal='center', vertical='center')
        cell_e1_text = "COLUMN B - STATUS"
        border_config()
    else:
        checkbox1.config(state="normal")
        checkbox2.config(state="normal")
        checkbox3.config(state="normal")
        checkbox4.config(state="normal")
        checkbox5.config(state="normal")
        checkbox6.config(state="normal")
        checkbox7.config(state="normal")
        checkbox8.config(state="normal")
        checkbox9.config(state="normal")

        entry_font.config(state="normal")

def toggle_menu():
    if menu_cores.winfo_ismapped():
        menu_cores.pack_forget()  # Oculta o menu de cores
    else:
        menu_cores.pack()

def border_config():
    global cell_border_left
    global cell_border_right
    global cell_border_top
    global cell_border_bottom

    global end_border_edges_bottom
    global end_border_edges_top

    global title_border_left
    global title_border_right
    global title_border_top
    global title_border_bottom

    if cell_border_left_config.get():
     cell_border_left = Side(style='thin', color=border_color)  # Borda fina preta no lado esquerdo
    else:
        cell_border_left = None

    if cell_border_right_config.get():
     cell_border_right = Side(style='thin', color=border_color)  # Borda fina preta no lado esquerdo
    else:
        cell_border_right = None
        
    if cell_border_top_config.get():
     cell_border_top = Side(style='thin', color=border_color)  # Borda fina preta no lado esquerdo
    else:
        cell_border_top = None

    if cell_border_bottom_config.get():
     cell_border_bottom = Side(style='thin', color=border_color)  # Borda fina preta no lado esquerdo
    else:
        cell_border_bottom = None

# ----------------------------------------------------------

    if title_border_left_config.get():
     title_border_left = Side(style='thin', color=border_color)  # Borda fina preta no lado esquerdo
    else:
        title_border_left = None
        
    if title_border_right_config.get():
     title_border_right = Side(style='thin', color=border_color)  # Borda fina preta no lado esquerdo
    else:
        title_border_right = None
        
    if title_border_top_config.get():
     title_border_top = Side(style='thin', color=border_color)  # Borda fina preta no lado esquerdo
    else:
        title_border_top = None

    if title_border_bottom_config.get():
     title_border_bottom = Side(style='thin', color=border_color)  # Borda fina preta no lado esquerdo
    else:
        title_border_bottom = None

# -------------------------------------------------------

    if end_border_edges_config.get():
     end_border_edges_top = Side(style='thin', color=border_color)
     end_border_edges_bottom = Side(style='thin', color=border_color)
    else:
        end_border_edges_top = None
        end_border_edges_bottom = None

def mudar_cor(cor):
    global border_color

    border_alt_color.configure(bg=cor)
    menu_cores.pack_forget()
    if cor == "#000000":
        border_alt_color.configure(fg="white")
    else:
        border_alt_color.configure(fg="black")

    if cor.startswith('#'):
        border_color = cor[1:]  # Remover '#' do início, se presente
    else:
        # Converter nome da cor para aRGB hexadecimal
        border_color = tk.colorchooser.color_to_rgb(cor).rgb[1:]

    border_config()

checkbox1 = tk.Checkbutton(root, text="Borda Direita da Célula ", variable=cell_border_right_config, command=border_config)
checkbox2 = tk.Checkbutton(root, text="Borda Esquerda da Célula ", variable=cell_border_left_config, command=border_config)
checkbox3 = tk.Checkbutton(root, text="Borda Superior da Célula ", variable=cell_border_top_config, command=border_config)
checkbox4 = tk.Checkbutton(root, text="Borda Inferior do Título ", variable=cell_border_bottom_config, command=border_config)
checkbox5 = tk.Checkbutton(root, text="Borda Direita do Título ", variable=title_border_right_config, command=border_config)
checkbox6 = tk.Checkbutton(root, text="Borda Esquerda do Título ", variable=title_border_left_config, command=border_config)
checkbox7 = tk.Checkbutton(root, text="Borda Superior do Título ", variable=title_border_top_config, command=border_config)
checkbox8 = tk.Checkbutton(root, text="Borda Inferior do Título ", variable=title_border_bottom_config, command=border_config)
checkbox9 = tk.Checkbutton(root, text="Borda das Extremidades", variable=end_border_edges_config, command=border_config)

def save_value():
    global input_column_value
    global output_column_value
    global font_size_e
    input_column_value = entry_input.get()
    output_column_value = entry_output.get()
    if check_font(entry_font.get()):
        font_size_e = entry_font.get()
    else:
        status_label.config(text="Font size should be between 8 and 18.", fg="red")


def process_file():

    global verified_sites
    global inaccessible_sites
    global accessible_sites
    global timeout_sites

    # Solicitar ao usuário que selecione o arquivo Excel
    filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])
    if filename:
        # Carregar o arquivo Excel e selecionar a planilha ativa
        timeout_sites = 0
        accessible_sites = 0
        inaccessible_sites = -1
        verified_sites = 0
        wb = load_workbook(filename)
        ws = wb.active

        # Definir o nome do novo arquivo Excel que será criado
        output_folder = os.path.expanduser("~/Documents/output_folder")
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        output_filename = os.path.join(output_folder, os.path.splitext(os.path.basename(filename))[0] + '_result2.xlsx')

        # Verificar se input_column_value está definido
        if input_column_value is not None:
            if output_column_value is not None:
                # Verificar se input_column_value é uma letra de coluna válida
                if len(input_column_value) == 1 and 'A' <= input_column_value.upper() <= 'Z':
                    if (len(output_column_value) == 1 and 'A' <= output_column_value.upper() <= 'Z' and output_column_value.upper() != input_column_value.upper()):
                        column_letter = input_column_value.upper()
                        output_column_letter = output_column_value.upper()
                        # Obter os valores de todas as células da coluna especificada e adicionar "http://" se necessário
                        sites = []
                        for cell in ws[column_letter]:
                            if cell.value is not None:
                                if cell.value.startswith("http://"):
                                    sites.append(cell.value)
                                else:
                                    sites.append("http://" + cell.value)
                
                        # Verificar se há sites antes de prosseguir
                        if sites:
                            
                            # Verificar os sites em paralelo
                            with ThreadPoolExecutor(max_workers=10) as executor:  # Executar até 10 verificações simultaneamente
                                results = list(executor.map(check_access, sites))
                                
                                # Adicionar os resultados na coluna E (ou 5) alinhados com os respectivos sites na coluna B
                                for i, result in enumerate(results, start=1):
                                    if i <= ws.max_row:
                                        ws.cell(row=i, column=column_index_from_string(output_column_letter)).value = result
                                        ws.cell(row=i, column=column_index_from_string(output_column_letter)).alignment = alignment_cell 
                                        
                                        # Atualizar a cor da fonte com base no resultado
                                        if result == "Accessible":
                                            accessible_sites += 1
                                            ws.cell(row=i, column=column_index_from_string(output_column_letter)).font = Font(size=font_size_e, color=font_color_acessible)
                                        elif result == "Inaccessible":
                                            inaccessible_sites += 1
                                            ws.cell(row=i, column=column_index_from_string(output_column_letter)).font = Font(size=font_size_e, color=font_color_inacessible)
                                        elif result == "Timeout Exceeded":
                                            timeout_sites += 1  # Incrementar timeout_sites se o tempo limite for excedido
                                            ws.cell(row=i, column=column_index_from_string(output_column_letter)).font = Font(size=font_size_e, color=font_color_timeout)

                            # Calcular o número total de sites verificados
                            verified_sites = accessible_sites + inaccessible_sites + timeout_sites

                            # Salvar o arquivo com os resultados
                            wb.save(output_filename)
                            status_label.config(text=f"Results saved in '{output_filename}'.")

                            # Escrever o texto definido na célula E1
                            ws[f'{output_column_value}1'].value = cell_e1_text
                            ws[f'{output_column_value}1'].font = Font(bold=True, size=11)
                            ws[f'{output_column_value}1'].alignment = Alignment(horizontal='center', vertical='center')
                            ws[f'{output_column_value}1'].border = Border(left=title_border_left, right=title_border_right, top=title_border_top, bottom=title_border_bottom)

                            # Aplicar bordas às células na coluna E
                            for i in range(2, verified_sites + 1):
                                cell = ws.cell(row=i, column=column_index_from_string(output_column_letter))
                                cell.border = Border(left=cell_border_left, right=cell_border_right, top=cell_border_top, bottom=cell_border_bottom)
                                
                            e2_cell = ws[f'{output_column_value}2']
                            e2_cell.border = Border(left=cell_border_left, right=cell_border_right, top=end_border_edges_top, bottom=cell_border_bottom)
                            ex_cell = ws[f'{output_column_value}{verified_sites + 1}']
                            ex_cell.border = Border(left=cell_border_left, right=cell_border_right, top=cell_border_top, bottom=end_border_edges_bottom)

                            # Definir a largura da coluna E
                            largura_coluna_B = ws.column_dimensions[column_letter].width
                            ws.column_dimensions[output_column_letter].width = largura_coluna_B
                            wb.save(output_filename)
                        else:
                            status_label.config(text=f"No sites found in column {column_letter}.")
                else:
                    status_label.config(text="Invalid column value. Please enter a letter from A to Z.")
        else:
            status_label.config(text="No column value entered. Please enter a column value first.")

# Criar a root principal

largura_root = 1000
altura_root = 400
root.geometry(f"{largura_root}x{altura_root}")

root.title("Site Verifier")

return_default_value = tk.BooleanVar()

# Criar um botão para abrir o arquivo Excel
open_button = tk.Button(root, text="Open Excel File", command=process_file)
open_button.pack(pady=10)

default_value = tk.Checkbutton(root, text="Retornar ao padrão", variable=return_default_value, command=default_return)
default_value.pack(anchor=tk.W)
default_value.select()
default_return()

checkbox1.pack(anchor=tk.W)
checkbox2.pack(anchor=tk.W)
checkbox3.pack(anchor=tk.W)
checkbox4.pack(anchor=tk.W)
checkbox5.pack(anchor=tk.W)
checkbox6.pack(anchor=tk.W)
checkbox7.pack(anchor=tk.W)
checkbox8.pack(anchor=tk.W)
checkbox9.pack(anchor=tk.W)

# Criando o botão inicial
border_alt_color = tk.Button(root, text="☰", fg="black", width=5, height=2, command=toggle_menu)
border_alt_color.pack(padx=10, pady=10)

border_alt_color.configure(bg="black")
border_alt_color.configure(fg="white")

# Criando o frame para o menu de cores
menu_cores = tk.Frame(root)
cores = ["#FF0000", "#00FF00", "#0000FF", "#FFFF00", "#FFA500", "#800080", "#000000", "#FFFFFF"]

frame_linha1 = tk.Frame(menu_cores)
frame_linha2 = tk.Frame(menu_cores)

for i in range(4):
    button_color = tk.Button(frame_linha1, bg=cores[i], width=5, height=2, command=lambda c=cores[i]: mudar_cor(c))
    button_color.pack(side=tk.LEFT, padx=5, pady=5)

for i in range(4, 8):
    button_color = tk.Button(frame_linha2, bg=cores[i], width=5, height=2, command=lambda c=cores[i]: mudar_cor(c))
    button_color.pack(side=tk.LEFT, padx=5, pady=5)

frame_linha1.pack()
frame_linha2.pack()

menu_cores.pack_forget()

root.bind("<Button-1>", root_click)

# Campo de entrada para aceitar um caractere de entrada
entry_input_column = tk.Frame(root)
entry_input_column.pack(pady=10)
entry_input_label = tk.Label(entry_input_column, text="Enter a input column letter (A-Z):")
entry_input_label.pack(side=tk.LEFT)
entry_input = tk.Entry(entry_input_column, width=5)
entry_input.pack(side=tk.LEFT)
entry_input.focus_set()

# Campo de entrada para aceitar um caractere de saída
entry_output_column = tk.Frame(root)
entry_output_column.pack(pady=10)
entry_output_label = tk.Label(entry_output_column, text="Enter a output column letter (A-Z):")
entry_output_label.pack(side=tk.LEFT)
entry_output = tk.Entry(entry_output_column, width=5)
entry_output.pack(side=tk.LEFT)

# Botão para salvar o caractere
save_button = tk.Button(text="Save", command=save_value)
save_button.pack(side=tk.LEFT)

# Rótulo para exibir o status
status_label = tk.Label(root, text="")
status_label.pack()

# Iniciar o loop de eventos
root.mainloop()
